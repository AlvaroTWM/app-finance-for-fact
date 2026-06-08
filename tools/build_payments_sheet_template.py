from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT_DIR = Path("/Users/alvaro.arambulo/Documents/React-loyalty-facturas-project/outputs/payments-template")
OUTPUT_FILE = OUTPUT_DIR / "plantilla_seguimiento_pagos_aliados.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="14532D")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FILL = PatternFill("solid", fgColor="DCFCE7")
TITLE_FONT = Font(color="14532D", bold=True, size=14)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def style_headers(sheet, headers):
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=col_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def write_title(sheet, title, description):
    sheet["A1"] = title
    sheet["A1"].font = TITLE_FONT
    sheet["A1"].fill = TITLE_FILL
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet["A2"] = description
    sheet["A2"].alignment = Alignment(wrap_text=True)


def write_rows(sheet, rows):
    for row_index, row_values in enumerate(rows, start=4):
        for col_index, value in enumerate(row_values, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.border = THIN_BORDER
            if isinstance(value, (int, float)) and col_index not in (4, 6):
                cell.number_format = '#,##0.00'


def set_widths(sheet, widths):
    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width


def add_status_validation(sheet, column_range, options):
    validation = DataValidation(
        type="list",
        formula1=f'"{",".join(options)}"',
        allow_blank=True,
    )
    sheet.add_data_validation(validation)
    validation.add(column_range)


def build_workbook():
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    aliados = workbook.create_sheet("Aliados")
    deudas = workbook.create_sheet("Deudas")
    cuotas = workbook.create_sheet("Cuotas")
    pagos = workbook.create_sheet("Pagos")

    write_title(
        aliados,
        "Aliados",
        "Ficha base de cada aliado. No registrar pagos aqui; solo informacion maestra.",
    )
    aliados_headers = [
        "aliado_id",
        "nombre_aliado",
        "telefono",
        "email",
        "estado",
    ]
    style_headers(aliados, aliados_headers)
    write_rows(
        aliados,
        [
            ["ALI-001", "Biggie Express", "0981-000000", "biggie@ejemplo.com", "activo"],
        ],
    )
    set_widths(aliados, {"A": 15, "B": 28, "C": 18, "D": 28, "E": 14})
    add_status_validation(aliados, "E4:E500", ["activo", "inactivo"])
    aliados.freeze_panes = "A4"

    write_title(
        deudas,
        "Deudas",
        "Resumen de la deuda por aliado. Aqui va el total, lo pagado acumulado y el saldo.",
    )
    deudas_headers = [
        "deuda_id",
        "aliado_id",
        "monto_total_deuda",
        "cantidad_cuotas",
        "monto_total_pagado",
        "saldo_pendiente",
        "estado_deuda",
        "fecha_acuerdo",
        "observaciones",
    ]
    style_headers(deudas, deudas_headers)
    write_rows(
        deudas,
        [
            [
                "DEU-001",
                "ALI-001",
                1200000,
                6,
                200000,
                1000000,
                "parcial",
                "2026-06-01",
                "Plan acordado con el aliado.",
            ],
        ],
    )
    set_widths(
        deudas,
        {"A": 15, "B": 15, "C": 20, "D": 18, "E": 20, "F": 18, "G": 16, "H": 16, "I": 32},
    )
    add_status_validation(deudas, "G4:G500", ["pendiente", "parcial", "pagado"])
    deudas.freeze_panes = "A4"

    write_title(
        cuotas,
        "Cuotas",
        "Detalle esperado de cada cuota por deuda. Cada fila representa una cuota del acuerdo.",
    )
    cuotas_headers = [
        "cuota_id",
        "deuda_id",
        "numero_cuota",
        "monto_cuota",
        "fecha_vencimiento",
        "estado_cuota",
        "fecha_pago",
        "monto_pagado",
    ]
    style_headers(cuotas, cuotas_headers)
    write_rows(
        cuotas,
        [
            ["CUO-001", "DEU-001", 1, 200000, "2026-06-10", "pagada", "2026-06-09", 200000],
            ["CUO-002", "DEU-001", 2, 200000, "2026-07-10", "pendiente", "", 0],
        ],
    )
    set_widths(cuotas, {"A": 15, "B": 15, "C": 15, "D": 18, "E": 18, "F": 16, "G": 16, "H": 16})
    add_status_validation(cuotas, "F4:F1000", ["pendiente", "pagada", "vencida"])
    cuotas.freeze_panes = "A4"

    write_title(
        pagos,
        "Pagos",
        "Historial real de pagos realizados. Cada fila representa un movimiento de pago.",
    )
    pagos_headers = [
        "pago_id",
        "aliado_id",
        "deuda_id",
        "cuota_id",
        "fecha_pago",
        "monto_pagado",
        "medio_pago",
        "comprobante",
        "comentario",
    ]
    style_headers(pagos, pagos_headers)
    write_rows(
        pagos,
        [
            [
                "PAG-001",
                "ALI-001",
                "DEU-001",
                "CUO-001",
                "2026-06-09",
                200000,
                "transferencia",
                "COMP-0001",
                "Pago registrado y conciliado.",
            ],
        ],
    )
    set_widths(
        pagos,
        {"A": 15, "B": 15, "C": 15, "D": 15, "E": 16, "F": 16, "G": 18, "H": 18, "I": 32},
    )
    pagos.freeze_panes = "A4"

    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = True

    return workbook


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook()
    workbook.save(OUTPUT_FILE)
    print(OUTPUT_FILE)


if __name__ == "__main__":
    main()
