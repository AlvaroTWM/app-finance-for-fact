from collections import Counter, defaultdict
import re
import unicodedata

from openpyxl import load_workbook


STOPWORDS = {
    "a","al","algo","algun","alguna","algunas","alguno","algunos","ante","antes","asi","aun",
    "bajo","bien","cada","casi","como","con","contra","cual","cuando","de","del","desde","donde",
    "dos","el","ella","ellas","ellos","en","entre","era","eramos","eran","eres","es","esa","esas",
    "ese","eso","esos","esta","estaba","estaban","estado","estamos","estan","estar","estas","este",
    "esto","estos","fue","fueron","ha","habia","hace","hacen","hacer","hacia","han","hasta","hay",
    "la","las","le","les","lo","los","mas","me","mi","mis","mucho","muy","necesita","ni","no","nos",
    "o","os","otra","otro","para","pero","poco","por","porque","que","quien","se","sea","segun","ser",
    "si","sin","sobre","son","su","sus","tal","tambien","te","tiene","tienen","todo","tras","tu","un",
    "una","uno","unos","usted","ya","e","u","y","cliente","caso","consulta","indica","comenta",
    "verifica","verificamos","solicitud","estimados","buenas","referencia","respecto","contacta",
    "favor","dias","dia","habiles","correo","mensaje","saludos","quedamos","estimado","estimada",
}


def normalize(text: str) -> str:
    text = str(text or "").lower().strip()
    text = "".join(
        ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn"
    )
    text = re.sub(r"[^\w\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def tokens(text: str):
    return [
        t for t in normalize(text).split()
        if t and t not in STOPWORDS and len(t) >= 3 and not t.isdigit()
    ]


wb = load_workbook("/Users/alvaro.arambulo/Downloads/Libro2.xlsx", read_only=True)
ws = wb[wb.sheetnames[0]]

rows = []
for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
    rows.append((idx, row[0]))

counter = Counter()
samples = defaultdict(list)

for _, value in rows:
    unique = set(tokens(value))
    for token in unique:
        counter[token] += 1
        if len(samples[token]) < 3:
            samples[token].append(str(value))

for token, count in counter.most_common(60):
    print(f"TOKEN\t{token}\t{count}")
    for sample in samples[token]:
        print(f"  SAMPLE\t{sample}")
