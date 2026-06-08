from collections import Counter, defaultdict
import re
import unicodedata

from openpyxl import load_workbook


def normalize(text: str) -> str:
    text = str(text or "").lower().strip()
    text = "".join(
        ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn"
    )
    text = re.sub(r"[^\w\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def tokenize(text: str):
    return [token for token in normalize(text).split() if token]


def has_phrase(text: str, phrases):
    return any(phrase in text for phrase in phrases)


def has_token(tokens, values):
    token_set = set(tokens)
    return any(value in token_set for value in values)


def has_token_prefix(tokens, prefixes):
    return any(any(token.startswith(prefix) for prefix in prefixes) for token in tokens)


def is_numeric_like(text: str) -> bool:
    compact = re.sub(r"[\s,.;:+\-_/()]+", "", text or "")
    return bool(compact) and compact.isdigit()


def classify(raw_value) -> str:
    text = normalize(raw_value)
    tokens = tokenize(raw_value)

    if not text:
        return "Sin descripcion"
    if is_numeric_like(text):
        return "Solo numero de caso"
    if (
        len(tokens) <= 8 and
        has_phrase(text, ["adjuntamos", "adjunto", "solicitud de prioridad", "se envia correo", "enviamos correo", "evidencia"])
    ) or (
        len(tokens) <= 6 and has_token(tokens, {"prioridad", "correo", "evidencia"})
    ):
        return "Seguimiento interno / adjuntos"
    if has_token(tokens, {"premio", "premios", "sorpresa", "camiseta", "camisetas", "entrada", "entradas", "canje", "qr"}):
        return "Premios / canjes"
    if has_token(tokens, {"desafio", "desafios", "mision", "misiones", "reto", "retos", "mundialista"}):
        return "Desafios / misiones / retos"
    if has_token_prefix(tokens, {"jubilad"}):
        return "Promocion jubilados"
    if has_token(tokens, {"upy", "upys"}) or has_phrase(text, ["sumar upys", "baja de upys", "bolsa de combustible", "subir al siguiente nivel", "perfil black", "nivel 4 al momento", "nivel actual"]):
        return "Upys / niveles / acciones"
    if has_phrase(text, ["dia de la madre", "beneficios para mama", "saldo promedio", "beneficio para mama", "para mama"]):
        return "Promocion dia de la madre / saldo promedio"
    if has_phrase(text, ["plazo", "8 dias", "ocho dias", "72 horas", "aguardar", "aun no paso", "aun no pasaron", "en revision", "dentro del plazo", "acreditacion", "se acreditara", "estado del reintegro", "no procesado", "se estara acreditando"]):
        return "Estado / plazo de reintegro"
    if has_phrase(text, ["tope", "limite", "monto minimo", "compra minima", "inferior", "no aplica", "no corresponde", "tarjeta debito", "pos de infonet", "pos de dinelco", "red infonet", "red dinelco", "comercio no se encuentra adherido", "no se encuentra adherido", "no contamos con una promocion especifica", "no califico para la promocion"]):
        return "Reintegro no aplica por condiciones"
    if has_phrase(text, ["reclama", "reclamo", "reclamar", "molest", "alterada", "acciones legales", "excepcion", "revision", "actualizacion del cupo", "anulada", "solicitamos su apoyo", "verificacion de un comercio aliado"]):
        return "Reclamo / revision de reintegro"
    if has_phrase(text, ["bases y condiciones", "byc", "promocion", "promociones", "comercios adheridos", "comercio adherido", "beneficios vigentes", "vigentes", "beneficios ueno", "tiene reintegro", "tiene reitegro", "locales adheridos"]):
        return "Consulta de promocion / ByC"
    if "reintegro" in text or "reintegros" in text:
        return "Consulta general de reintegros"
    return "Otros"


wb = load_workbook("/Users/alvaro.arambulo/Downloads/Libro2.xlsx", read_only=True)
ws = wb[wb.sheetnames[0]]

counts = Counter()
samples = defaultdict(list)

for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
    value = row[0]
    group = classify(value)
    counts[group] += 1
    if len(samples[group]) < 5:
        samples[group].append((idx, str(value)))

for group, count in counts.most_common():
    print(f"GROUP\t{group}\t{count}")
    for idx, sample in samples[group]:
        print(f"  SAMPLE\t{idx}\t{sample}")
